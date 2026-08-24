from fastapi import APIRouter, Request, Form, File, UploadFile, HTTPException
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel
import db
import ai
import re
import json
import auth
import io
import os
import shutil
import config_manager
from datetime import datetime

router = APIRouter()

class TestChatMessage(BaseModel):
    phone: str
    message: str
    profile_name: str = "Tester"

class ManagerMessagePayload(BaseModel):
    phone: str
    message: str


@router.get("/dashboard")
def get_dashboard(request: Request):
    if not auth.verify_auth(request):
        return RedirectResponse(url="/login")
    return FileResponse("templates/dashboard.html")

@router.get("/test-chat")
def get_test_chat(request: Request):
    if not auth.verify_auth(request):
        return RedirectResponse(url="/login")
    return FileResponse("templates/test_chat.html")

@router.post("/api/test-chat")
def api_test_chat(msg: TestChatMessage, request: Request):
    """Bypasses Meta API and interacts directly with AI for testing."""
    auth.require_auth(request)
    incoming_msg = msg.message.strip()
    if not incoming_msg:
        return {"error": "Message is empty"}
        
    db.log_chat_message(msg.phone, "inbound", incoming_msg)
    
    ai_response = ai.get_ai_response(msg.phone, msg.profile_name)
    reply_text = ai_response
    image_file = None
    
    submit_match = re.search(r'\[LEAD_SUBMIT:\s*(\{.*?\})\s*\]', ai_response, re.DOTALL)
    status_match = "[LEAD_STATUS_CHECK]" in ai_response
    image_match = re.search(r'\[IMAGE:\s*(.+?)\s*\]', ai_response)
    
    if image_match:
        image_file = image_match.group(1).strip()
        reply_text = re.sub(r'\[IMAGE:\s*.+?\s*\]', '', reply_text).strip()

    if submit_match:
        try:
            lead_data = json.loads(submit_match.group(1))
            # Validate required fields
            if not lead_data.get("name") or not lead_data.get("product"):
                raise ValueError("Missing required lead fields: name and product")
            lead_id = db.create_lead(
                phone=msg.phone,
                name=lead_data.get("name", "Unknown")[:200],
                company=lead_data.get("company", "Individual")[:200],
                email=lead_data.get("email", "")[:200],
                location=lead_data.get("location", "Unknown")[:200],
                product_interest=lead_data.get("product", "Unknown")[:200],
                quantity=lead_data.get("quantity", "Unknown")[:100],
                requirements=f"Captured via AI chatbot. Qty: {lead_data.get('quantity')}. Loc: {lead_data.get('location')}."
            )
            cleaned_text = re.sub(r'\[LEAD_SUBMIT:\s*\{.*?\}\s*\]', '', ai_response, flags=re.DOTALL).strip()
            success_msg = "🎉 *Inquiry Submitted Successfully!*\n\nOur sales representatives are reviewing your requirements and will reach out shortly."
            reply_text = f"{cleaned_text}\n\n{success_msg}" if cleaned_text else success_msg
        except Exception:
            reply_text = "I encountered an error submitting your quote request. Please try again."
            
    elif status_match:
        lead = db.get_lead_by_phone(msg.phone)
        cleaned_text = ai_response.replace("[LEAD_STATUS_CHECK]", "").strip()
        if lead:
            reply_text = f"{cleaned_text}\n\n📄 *Inquiry #{lead['id']} Status:* {lead['status']}" if cleaned_text else f"📄 *Inquiry #{lead['id']} Status:* {lead['status']}"
        else:
            reply_text = f"{cleaned_text}\n\n❌ No active inquiry found." if cleaned_text else "❌ No active inquiry found."

    reply_text = reply_text.replace("**", "*")
    reply_text = re.sub(r'\n{3,}', '\n\n', reply_text).strip()
    db.log_chat_message(msg.phone, "outbound", reply_text)
    
    return {"reply": reply_text, "image": image_file}

@router.get("/api/leads")
def get_leads_api(request: Request, status: str = None, search: str = None):
    auth.require_auth(request)
    return db.get_leads(status_filter=status, search_query=search)

@router.get("/api/leads/export")
def export_leads_excel(request: Request):
    auth.require_auth(request)
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    leads = db.get_leads()

    wb = Workbook()
    ws = wb.active
    ws.title = "KDI Leads"

    # ── Styles ────────────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="C0521F")   # KDI orange
    HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ALT_ROW_FILL  = PatternFill("solid", fgColor="F5F5F5")   # light grey
    WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")
    CELL_FONT     = Font(name="Calibri", size=10)
    CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGN    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    STATUS_COLORS = {
        "New":       "1A73E8",   # blue
        "Contacted": "F9AB00",   # amber
        "Quoted":    "9334E6",   # purple
        "Won":       "1E8E3E",   # green
        "Lost":      "D93025",   # red
        "Partial":   "FA7B17",   # orange
    }

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = [
        "#", "Status", "Name", "Company", "Phone", "Email", "Location",
        "Product Required", "Quantity", "Requirements / Notes",
        "Inquiry Date", "Days Open"
    ]
    ws.append(headers)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER_ALIGN

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"   # freeze header row

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_num, lead in enumerate(leads, start=1):
        status = lead.get("status", "")

        # Robust phone extraction checking multiple possible schema keys
        raw_phone = str(lead.get("phone") or lead.get("phone_number") or lead.get("mobile") or "").strip()
        if raw_phone and not raw_phone.startswith("+"):
            formatted_phone = f"+{raw_phone}"
        else:
            formatted_phone = raw_phone

        # Inquiry date — just the date, no time clutter
        def fmt_date(val):
            if not val:
                return ""
            try:
                return datetime.fromisoformat(val.replace("Z", "+00:00")).strftime("%d %b %Y")
            except Exception:
                return str(val)

        # Days open — how many days since the inquiry was submitted
        def days_open(val):
            if not val:
                return ""
            try:
                created = datetime.fromisoformat(val.replace("Z", "+00:00"))
                delta = (datetime.utcnow().replace(tzinfo=created.tzinfo) - created).days
                return delta
            except Exception:
                return ""

        row_data = [
            row_num,
            status,
            lead.get("name", ""),
            lead.get("company", ""),
            formatted_phone,
            lead.get("email", ""),
            lead.get("location", ""),
            lead.get("product_interest", ""),
            lead.get("quantity", ""),
            lead.get("requirements", ""),
            fmt_date(lead.get("created_at")),
            days_open(lead.get("created_at")),
        ]
        ws.append(row_data)

        excel_row = row_num + 1   # +1 because header is row 1
        row_fill = ALT_ROW_FILL if row_num % 2 == 0 else WHITE_FILL

        # Columns centred: #(1), Status(2), Phone(5), Inquiry Date(11), Days Open(12)
        CENTRE_COLS = {1, 2, 5, 11, 12}

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border    = BORDER
            cell.alignment = CENTER_ALIGN if col_idx in CENTRE_COLS else LEFT_ALIGN
            cell.font      = CELL_FONT

            # Ensure Phone column (5) is formatted as text so Excel displays it properly
            if col_idx == 5:
                cell.number_format = '@'

            # Status column (col 2) gets colour-coded bold text
            if col_idx == 2 and status in STATUS_COLORS:
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color=STATUS_COLORS[status])
                cell.fill = WHITE_FILL
            # Days Open (col 12) — highlight if > 3 days with a warm tint
            elif col_idx == 12:
                days = row_data[11]
                if isinstance(days, int) and days > 3:
                    cell.fill = PatternFill("solid", fgColor="FFF3CD")   # soft amber
                    cell.font = Font(name="Calibri", size=10, bold=True, color="7B5800")
                else:
                    cell.fill = row_fill
            else:
                cell.fill = row_fill

        ws.row_dimensions[excel_row].height = 18

    # ── Auto-fit column widths ─────────────────────────────────────────────────
    MIN_WIDTH = 10
    MAX_WIDTH = 45
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        adjusted = min(MAX_WIDTH, max(MIN_WIDTH, max_len + 3))
        ws.column_dimensions[col_letter].width = adjusted

    # ── Stream response ────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"KDI_Leads_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.patch("/api/leads/{lead_id}/status")
async def update_lead_status_api(lead_id: int, request: Request):
    auth.require_auth(request)
    payload = await request.json()
    status = payload.get("status")
    ALLOWED_STATUSES = {"New", "Contacted", "Quoted", "Won", "Lost", "Partial"}
    if status not in ALLOWED_STATUSES:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {', '.join(sorted(list(ALLOWED_STATUSES)))}")
    db.update_lead_status(lead_id, status)
    return {"success": True, "lead_id": lead_id, "status": status}

@router.put("/api/leads/{lead_id}")
async def update_lead_details_api(lead_id: int, request: Request):
    """Updates full lead details (name, company, email, location, product, quantity, requirements)."""
    auth.require_auth(request)
    data = await request.json()
    db.update_lead_details(lead_id, data)
    return {"success": True, "lead_id": lead_id, "message": "Lead details updated successfully."}


@router.delete("/api/leads/clear-all")
def clear_all_leads_api(request: Request):
    """Deletes all leads and chat history from Supabase database."""
    auth.require_auth(request)
    db.request_supabase("leads", "DELETE", params={"id": "gt.0"})
    db.request_supabase("chat_history", "DELETE", params={"id": "gt.0"})
    return {"success": True, "message": "All leads and chat history have been cleared."}

@router.delete("/api/leads/{lead_id}")
def delete_lead_api(lead_id: int, request: Request):
    """Deletes a single lead and its chat history by lead ID."""
    auth.require_auth(request)
    db.delete_lead(lead_id)
    return {"success": True, "lead_id": lead_id, "message": "Lead deleted successfully."}

@router.get("/api/leads/{phone}/history")
def get_lead_history_api(phone: str, request: Request):
    auth.require_auth(request)
    return db.get_chat_history(phone)

@router.post("/api/leads/send-message")
async def send_manager_message_api(payload: ManagerMessagePayload, request: Request):
    """Allows a manager/admin to send a direct WhatsApp reply to a customer."""
    auth.require_auth(request)
    msg_text = payload.message.strip()
    if not msg_text:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Message content cannot be empty.")
    
    phone = payload.phone.strip()
    if not phone:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Phone number is required.")

    # 1. Send via WhatsApp Meta Cloud API
    from routes.whatsapp import send_whatsapp_message
    try:
        send_whatsapp_message(to_phone=phone, text=msg_text)
    except RuntimeError as re_err:
        from fastapi import HTTPException
        if "24H_WINDOW_EXPIRED" in str(re_err):
            raise HTTPException(
                status_code=400,
                detail="Meta 24-hour messaging window expired. Click 'Chat on WhatsApp Web' button above to send a direct message."
            )
        raise HTTPException(status_code=400, detail=str(re_err))
    except Exception as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=f"Failed to send WhatsApp message: {str(e)}")

    # 2. Log in chat history DB as outbound message
    db.log_chat_message(phone, "outbound", msg_text)

    # 3. Auto-update lead status to "Contacted" if it was "New" or "Partial"
    lead = db.get_lead_by_phone(phone)
    if lead and lead.get("status") in ["New", "Partial"]:
        db.update_lead_status(lead["id"], "Contacted")

    return {"success": True, "message": "Message sent via WhatsApp successfully."}

@router.get("/api/visitors")
def get_visitors_api(request: Request):
    """Returns a list of chat contacts who are NOT in the leads table."""
    auth.require_auth(request)
    return db.get_visitor_chats()

@router.delete("/api/visitors/clear-all")
def clear_all_visitors_api(request: Request):
    """Deletes all non-lead visitor chats from chat history."""
    auth.require_auth(request)
    deleted = db.clear_visitor_chats()
    return {"success": True, "deleted": deleted}

@router.delete("/api/visitors/{phone}")
def delete_visitor_api(phone: str, request: Request):
    """Deletes a single visitor's chat history (non-lead phones only)."""
    auth.require_auth(request)
    # Guard: never delete chat history for a phone that now has a lead record
    existing_lead = db.get_lead_by_phone(phone)
    if existing_lead:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=400,
            detail="This phone now has a lead record. Delete the lead from the Leads Inbox instead."
        )
    db.delete_visitor_chats(phone)
    return {"success": True, "phone": phone}

@router.post("/api/visitors/{phone}/convert")
def convert_visitor_api(phone: str, request: Request):
    """Converts a visitor chat into a tracked sales lead in the Leads Inbox."""
    auth.require_auth(request)

    # Carry over the visitor's most recent inbound message as context
    context = "Converted from visitor chat."
    chats = db.get_chat_history(phone, limit=20)
    inbound_msgs = [c for c in chats if c.get("direction") == "inbound"]
    if inbound_msgs:
        body = (inbound_msgs[-1].get("body") or "").strip()
        if body:
            context = f"Converted from visitor chat. Last message: {body[:200]}"

    # Use AI to extract lead info from chat history
    extracted_details = {}
    try:
        from ai import extract_lead_info_from_history
        extracted_details = extract_lead_info_from_history(chats)
    except Exception as e:
        logger.error(f"Error extracting lead info from chat history: {e}")

    result = db.convert_visitor_to_lead(phone, context, extracted_details=extracted_details)
    if result == "exists":
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="This phone already has a lead record in the Leads Inbox.")
    if result != "created":
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="Failed to convert visitor to lead. Please check the Supabase connection.")
    return {"success": True, "phone": phone, "message": "Visitor converted to lead successfully.", "details": extracted_details}

@router.get("/api/outbound-messages")
def get_outbound_messages_api(request: Request):
    """Returns list of manager direct outbound messages."""
    auth.require_auth(request)
    return db.get_all_outbound_messages()

@router.post("/api/messages/send-media")
async def send_manager_media_api(
    request: Request,
    phone: str = Form(...),
    message: str = Form(""),
    file: UploadFile = File(None)
):
    """Sends a direct WhatsApp text message, image, or PDF document file to a contact."""
    auth.require_auth(request)
    phone = phone.strip()
    message_text = message.strip()
    
    if not phone:
        raise HTTPException(status_code=400, detail="Phone number is required.")
        
    if not message_text and not file:
        raise HTTPException(status_code=400, detail="Please enter a message or attach a file.")
        
    file_url = None
    file_type = None
    saved_filename = None
    
    if file and file.filename:
        os.makedirs("static/uploads", exist_ok=True)
        # Clean filename
        clean_name = re.sub(r'[^a-zA-Z0-9_.-]', '_', file.filename)
        saved_filename = f"{int(datetime.utcnow().timestamp())}_{clean_name}"
        save_path = os.path.join("static", "uploads", saved_filename)
        
        with open(save_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        base_url = os.environ.get("BASE_URL", "https://whatsapp-bot-4ukk.onrender.com").rstrip("/")
        file_url = f"{base_url}/static/uploads/{saved_filename}"
        
        ext = os.path.splitext(clean_name)[1].lower()
        if ext in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
            file_type = "image"
        else:
            file_type = "document"

    from routes.whatsapp import send_whatsapp_message, send_whatsapp_document
    
    caption_text = message_text
    
    try:
        if file_type == "image":
            send_whatsapp_message(to_phone=phone, text=caption_text if message_text else "", image_url=file_url)
            log_body = f"{caption_text}\n🖼️ [Attached Image: {saved_filename}]".strip()
        elif file_type == "document":
            send_whatsapp_document(to_phone=phone, document_url=file_url, filename=file.filename, caption=caption_text if message_text else "")
            log_body = f"{caption_text}\n📄 [Attached Document: {file.filename}]".strip()
        else:
            send_whatsapp_message(to_phone=phone, text=caption_text)
            log_body = caption_text
    except Exception as e:
        logger.error(f"Error sending WhatsApp media message: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send WhatsApp message: {str(e)}")

    db.log_chat_message(phone, "outbound", log_body)

    # Auto update lead status if lead exists
    lead = db.get_lead_by_phone(phone)
    if lead and lead.get("status") in ["New", "Partial"]:
        db.update_lead_status(lead["id"], "Contacted")

    return {"success": True, "message": "Outbound message sent via WhatsApp successfully.", "file_url": file_url}

@router.post("/api/settings/upload-catalogue")
async def upload_catalogue_pdf_api(request: Request, file: UploadFile = File(...)):
    """Uploads a new product catalogue PDF file, updating the active PDF."""
    auth.require_auth(request)
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No PDF file uploaded.")
        
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF document (.pdf).")
        
    os.makedirs("data/raw_catalogues", exist_ok=True)
    os.makedirs("static/uploads", exist_ok=True)
    
    # Save to canonical catalogue location
    target_path = os.path.join("data", "raw_catalogues", "CATALOUGE.pdf")
    with open(target_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # Also save copy to static/uploads for direct previewing
    static_copy = os.path.join("static", "uploads", "CATALOUGE.pdf")
    shutil.copyfile(target_path, static_copy)

    # Update settings config with filename & timestamp
    cfg = config_manager.get_config()
    cfg["catalogue_filename"] = file.filename
    cfg["catalogue_updated_at"] = datetime.utcnow().isoformat() + "Z"
    config_manager.save_config(cfg)
    
    return {"success": True, "message": f"Catalogue PDF updated successfully ({file.filename})", "filename": file.filename, "url": "/catalogue/CATALOUGE.pdf"}


@router.get("/api/dashboard/stats")
def get_stats_api(request: Request):
    auth.require_auth(request)
    leads = db.get_leads()
    
    total_leads = len(leads)
    new_leads = sum(1 for l in leads if l.get("status") == "New")
    contacted_leads = sum(1 for l in leads if l.get("status") == "Contacted")
    quoted_leads = sum(1 for l in leads if l.get("status") == "Quoted")
    won_leads = sum(1 for l in leads if l.get("status") == "Won")
    lost_leads = sum(1 for l in leads if l.get("status") == "Lost")
    
    # Category / Product interest distribution for charts
    categories = {}
    
    # Pre-populate product name to category mapping dynamically
    prod_to_cat = {}
    try:
        # Load categories from catalog products
        all_prods = db.get_all_products()
        for p in all_prods:
            prod_to_cat[p["name"]] = p["category"]
    except Exception:
        pass
        
    for l in leads:
        prod = l.get("product_interest")
        if not prod:
            continue

        cat = prod_to_cat.get(prod)
        if not cat:
            prod_lower = prod.lower()
            if "control" in prod_lower:
                cat = "Control Cables"
            elif "aerial" in prod_lower or "bunched" in prod_lower or "abc" in prod_lower:
                cat = "Aerial Bunched Cable"
            elif "submersible" in prod_lower or "flexible" in prod_lower or "cord" in prod_lower or "trailing" in prod_lower or "rubber" in prod_lower:
                cat = "Rubber Cable"
            elif "thermocouple" in prod_lower or "instrumentation" in prod_lower or "compensating" in prod_lower or "shielded" in prod_lower:
                cat = "Instrumentation Wires"
            elif "house" in prod_lower or "fr" in prod_lower or "wire" in prod_lower or "triple" in prod_lower:
                cat = "House Wires"
            else:
                cat = "Power Cables"

        categories[cat] = categories.get(cat, 0) + 1

    return {
        "total_leads": total_leads,
        "new_leads": new_leads,
        "contacted_leads": contacted_leads,
        "quoted_leads": quoted_leads,
        "won_leads": won_leads,
        "lost_leads": lost_leads,
        "category_distribution": categories
    }

@router.get("/api/products")
def get_products_api(request: Request):
    auth.require_auth(request)
    return db.get_all_products()

@router.patch("/api/products/{product_name}")
async def update_product_api(product_name: str, request: Request):
    auth.require_auth(request)
    payload = await request.json()
    price = payload.get("price")
    stock_status = payload.get("stock_status")
    if price is not None and (not isinstance(price, (int, float)) or price < 0):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Price must be a non-negative number.")
    db.update_product_price_and_stock(product_name, price, stock_status)
    return {"success": True, "product": product_name}

@router.post("/api/products")
async def create_product_api(request: Request):
    """Creates a new product in the catalog."""
    auth.require_auth(request)
    payload = await request.json()
    
    name = (payload.get("name") or "").strip()
    category = (payload.get("category") or "").strip()
    conductor = (payload.get("conductor") or "").strip()
    size = (payload.get("size") or "").strip()
    core = payload.get("core")
    insulation = (payload.get("insulation") or "XLPE").strip()
    price = payload.get("price_per_meter")
    stock_status = (payload.get("stock_status") or "In Stock").strip()
    specifications = (payload.get("specifications") or "").strip()
    
    # Validate required fields
    if not name or not category or not conductor or not size:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Product name, category, conductor, and size are required.")
    
    if price is not None and (not isinstance(price, (int, float)) or price < 0):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Price must be a non-negative number.")
    
    product_data = {
        "name": name[:200],
        "category": category[:100],
        "conductor": conductor[:100],
        "size": size[:50],
        "core": int(float(core)) if core else 1,
        "insulation": insulation[:100],
        "price_per_meter": float(price) if price else 0,
        "stock_status": stock_status,
        "specifications": specifications[:500]
    }
    
    result = db.create_product(product_data)
    if result == "exists":
        from fastapi import HTTPException
        raise HTTPException(status_code=409, detail=f"A product with the name '{name}' already exists. Use the edit feature to update it.")
    elif result == "created":
        return {"success": True, "product": name, "action": "created"}
    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="Failed to create product. Please check Supabase connection.")

@router.delete("/api/products/{product_name}")
async def delete_product_api(product_name: str, request: Request):
    """Deletes a product from the catalog."""
    auth.require_auth(request)
    
    import urllib.parse
    decoded_name = urllib.parse.unquote(product_name)
    
    existing = db.get_product_by_id(decoded_name)
    if not existing:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=f"Product '{decoded_name}' not found.")
    
    db.request_supabase("products", "DELETE", params={"name": f"eq.{decoded_name}"})
    return {"success": True, "product": decoded_name, "message": "Product deleted successfully."}

@router.get("/api/settings")
def get_settings_api(request: Request):
    auth.require_auth(request)
    return config_manager.get_config()

@router.put("/api/settings")
async def update_settings_api(request: Request):
    auth.require_auth(request)
    payload = await request.json()
    success = config_manager.save_config(payload)
    if success:
        return {"success": True}
    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="Failed to save configuration.")

# ── Template Management API ───────────────────────────────

from logger import get_logger
logger = get_logger(__name__)

def _sync_meta_templates_into_config():
    from routes.whatsapp import get_meta_templates
    meta_templates = get_meta_templates()
    if not meta_templates:
        return config_manager.get_templates(), 0, 0

    local_templates = config_manager.get_templates()
    local_by_name = {tpl["name"]: tpl for tpl in local_templates}
    updated_count = 0
    imported_count = 0

    for mt in meta_templates:
        mt_name = mt.get("name", "").strip()
        if not mt_name:
            continue
        
        status = mt.get("status", "PENDING")
        meta_id = mt.get("id")
        rejection_reason = mt.get("rejected_reason") if status == "REJECTED" else None
        
        if mt_name in local_by_name:
            tpl = local_by_name[mt_name]
            changed = False
            if tpl.get("meta_status") != status:
                tpl["meta_status"] = status
                changed = True
            if tpl.get("meta_template_id") != meta_id:
                tpl["meta_template_id"] = meta_id
                changed = True
            if tpl.get("meta_rejection_reason") != rejection_reason:
                tpl["meta_rejection_reason"] = rejection_reason
                changed = True
            if changed:
                tpl["updated_at"] = datetime.utcnow().isoformat() + "Z"
                config_manager.save_template(tpl)
                updated_count += 1
        else:
            # Import new template from Meta into local store
            components = mt.get("components", [])
            header = None
            body = ""
            footer = ""
            buttons = []
            for comp in components:
                ctype = comp.get("type", "").upper()
                if ctype == "HEADER":
                    hformat = comp.get("format", "TEXT").lower()
                    hval = comp.get("text", "")
                    header = {"type": hformat, "value": hval}
                elif ctype == "BODY":
                    body = comp.get("text", "")
                elif ctype == "FOOTER":
                    footer = comp.get("text", "")
                elif ctype == "BUTTONS":
                    for b in comp.get("buttons", []):
                        buttons.append({
                            "type": b.get("type", "QUICK_REPLY").lower(),
                            "text": b.get("text", "")
                        })

            new_tpl = config_manager.create_template_obj(
                name=mt_name,
                category=mt.get("category", "MARKETING"),
                language=mt.get("language", "en"),
                header=header,
                body=body,
                footer=footer,
                buttons=buttons
            )
            new_tpl["meta_status"] = status
            new_tpl["meta_template_id"] = meta_id
            new_tpl["meta_rejection_reason"] = rejection_reason
            config_manager.save_template(new_tpl)
            local_by_name[mt_name] = new_tpl
            imported_count += 1

    return config_manager.get_templates(), updated_count, imported_count

@router.get("/api/templates")
def get_templates_api(request: Request):
    """Returns all saved message templates. Auto-syncs from Meta if local cache is empty."""
    auth.require_auth(request)
    templates = config_manager.get_templates()
    if not templates:
        try:
            templates, _, _ = _sync_meta_templates_into_config()
        except Exception as e:
            logger.error(f"Auto-sync error on get_templates: {e}")
    return {"templates": templates}

@router.post("/api/templates")
async def create_template_api(request: Request):
    """Creates a new message template and optionally submits to Meta for approval."""
    auth.require_auth(request)
    payload = await request.json()

    name = payload.get("name", "").strip().lower().replace(" ", "_")
    name = re.sub(r'[^a-z0-9_]', '', name)
    if not name:
        raise HTTPException(status_code=400, detail="Template name is required (lowercase, underscores only).")

    category = payload.get("category", "MARKETING")
    language = payload.get("language", "en")
    header = payload.get("header")  # { type, content }
    body = payload.get("body", "")
    footer = payload.get("footer", "")
    buttons = payload.get("buttons", [])
    submit_to_meta = payload.get("submit_to_meta", False)

    if not body:
        raise HTTPException(status_code=400, detail="Template body text is required.")

    # Create local template object
    template = config_manager.create_template_obj(
        name=name, category=category, language=language,
        header=header, body=body, footer=footer, buttons=buttons
    )

    # Optionally submit to Meta
    if submit_to_meta:
        try:
            from routes.whatsapp import build_meta_components, create_meta_template
            components = build_meta_components(header, body, footer, buttons)
            result = create_meta_template(name, category, language, components)
            template["meta_template_id"] = result.get("id")
            template["meta_status"] = result.get("status", "PENDING")
        except Exception as e:
            logger.error(f"Meta template submission failed: {e}")
            template["meta_status"] = "LOCAL"
            # Still save locally, just note it wasn't submitted

    config_manager.save_template(template)
    return {"success": True, "template": template}

@router.delete("/api/templates/{template_id}")
def delete_template_api(template_id: str, request: Request):
    """Deletes a template locally and optionally from Meta."""
    auth.require_auth(request)
    templates = config_manager.get_templates()
    target = next((t for t in templates if t["id"] == template_id), None)
    if not target:
        raise HTTPException(status_code=404, detail="Template not found.")

    # Delete from Meta if it was submitted
    if target.get("meta_template_id"):
        try:
            from routes.whatsapp import delete_meta_template
            delete_meta_template(target["name"])
        except Exception as e:
            logger.error(f"Failed to delete Meta template: {e}")

    config_manager.delete_template(template_id)
    return {"success": True}

@router.post("/api/templates/sync")
def sync_templates_api(request: Request):
    """Syncs local template statuses with Meta and imports missing templates."""
    auth.require_auth(request)
    try:
        templates, updated_count, imported_count = _sync_meta_templates_into_config()
        return {"success": True, "synced": updated_count, "imported": imported_count, "total": len(templates)}
    except Exception as e:
        logger.error(f"Failed to sync Meta templates: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch Meta templates: {str(e)}")

@router.post("/api/templates/{template_id}/send")
async def send_template_api(template_id: str, request: Request):
    """Sends a template message to a single phone number."""
    auth.require_auth(request)
    payload = await request.json()
    phone = payload.get("phone", "").strip().replace("+", "")
    phone = re.sub(r'[^0-9]', '', phone)
    variables = payload.get("variables", {})  # { "1": "value1", "2": "value2" }

    if not phone:
        raise HTTPException(status_code=400, detail="Phone number is required.")

    templates = config_manager.get_templates()
    template = next((t for t in templates if t["id"] == template_id), None)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found.")

    # Auto-fill missing variables (e.g. {{1}} -> lead name)
    body_var_nums = re.findall(r'\{\{(\d+)\}\}', template.get("body", ""))
    if body_var_nums:
        lead = db.get_lead_by_phone(phone)
        contact_name = (lead.get("name") if lead and lead.get("name") else "") or "Customer"
        for vn in body_var_nums:
            if not variables.get(vn):
                variables[vn] = contact_name if vn == "1" else f"Value {vn}"

    # Build send-time components (fill in variable parameters)
    send_components = []
    if variables:
        params = []
        for key in sorted(variables.keys(), key=lambda x: int(x)):
            params.append({"type": "text", "text": str(variables[key])})
        if params:
            send_components.append({"type": "body", "parameters": params})

    # If header is media, add header component
    if template.get("header") and template["header"].get("type") in ["image", "document", "video"]:
        h_type = template["header"]["type"]
        h_content = template["header"].get("content", "")
        if h_content:
            header_param = {"type": h_type}
            header_param[h_type] = {"link": h_content}
            send_components.append({"type": "header", "parameters": [header_param]})

    try:
        from routes.whatsapp import send_whatsapp_template
        result = send_whatsapp_template(
            to_phone=phone,
            template_name=template["name"],
            language_code=template.get("language", "en"),
            components=send_components if send_components else None
        )
        # Log the outbound message
        body_text = template.get("body", "")
        for k, v in variables.items():
            body_text = body_text.replace(f"{{{{{k}}}}}", str(v))
        db.log_chat_message(phone, "outbound", f"[Template: {template['name']}] {body_text}")

        # Auto-update lead status
        lead = db.get_lead_by_phone(phone)
        if lead and lead.get("status") in ["New", "Partial"]:
            db.update_lead_status(lead["id"], "Contacted")

        return {"success": True, "message": f"Template sent to +{phone}"}
    except Exception as e:
        logger.error(f"Error sending template: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send template: {str(e)}")

# ── Mass Broadcast API ────────────────────────────────────

@router.post("/api/broadcast")
async def send_broadcast_api(request: Request):
    """Sends a template message to multiple phone numbers (mass broadcast)."""
    auth.require_auth(request)
    payload = await request.json()

    template_id = payload.get("template_id", "")
    phones = payload.get("phones", [])  # list of phone strings
    variables = payload.get("variables", {})

    if not template_id:
        raise HTTPException(status_code=400, detail="Template ID is required.")
    if not phones or len(phones) == 0:
        raise HTTPException(status_code=400, detail="At least one recipient phone is required.")

    templates = config_manager.get_templates()
    template = next((t for t in templates if t["id"] == template_id), None)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found.")

    # Build send-time components
    send_components = []
    if variables:
        params = []
        for key in sorted(variables.keys(), key=lambda x: int(x)):
            params.append({"type": "text", "text": str(variables[key])})
        if params:
            send_components.append({"type": "body", "parameters": params})

    if template.get("header") and template["header"].get("type") in ["image", "document", "video"]:
        h_type = template["header"]["type"]
        h_content = template["header"].get("content", "")
        if h_content:
            header_param = {"type": h_type}
            header_param[h_type] = {"link": h_content}
            send_components.append({"type": "header", "parameters": [header_param]})

    from routes.whatsapp import send_whatsapp_template
    import time

    results = {"total": len(phones), "sent": 0, "failed": 0, "errors": []}

    for phone in phones:
        clean_phone = re.sub(r'[^0-9]', '', str(phone).strip())
        if not clean_phone:
            results["failed"] += 1
            results["errors"].append({"phone": phone, "error": "Invalid phone number"})
            continue
        try:
            send_whatsapp_template(
                to_phone=clean_phone,
                template_name=template["name"],
                language_code=template.get("language", "en"),
                components=send_components if send_components else None
            )
            # Log outbound
            body_text = template.get("body", "")
            for k, v in variables.items():
                body_text = body_text.replace(f"{{{{{k}}}}}", str(v))
            db.log_chat_message(clean_phone, "outbound", f"[Broadcast: {template['name']}] {body_text}")
            results["sent"] += 1
            # Rate limiting: ~10 messages per second max for Meta
            time.sleep(0.15)
        except Exception as e:
            results["failed"] += 1
            results["errors"].append({"phone": clean_phone, "error": str(e)})
            logger.error(f"Broadcast send error to {clean_phone}: {e}")

    # Log broadcast in history
    broadcast_entry = {
        "id": f"bc_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
        "template_name": template["name"],
        "template_id": template_id,
        "total": results["total"],
        "sent": results["sent"],
        "failed": results["failed"],
        "created_at": datetime.utcnow().isoformat()
    }
    config_manager.add_broadcast_entry(broadcast_entry)

    return {"success": True, "results": results, "broadcast": broadcast_entry}

@router.get("/api/broadcast/history")
def get_broadcast_history_api(request: Request):
    """Returns broadcast history."""
    auth.require_auth(request)
    return {"history": config_manager.get_broadcast_history()}

